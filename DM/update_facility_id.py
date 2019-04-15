import numpy as np
import pandas as pd
import cx_Oracle as oracle
import os

from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter
import time

import conf.acct_oracle as acct_oracle
from db_connect.oracle_db import UseOracleDB
from tool.df_compare import has_gap
import tool.oracle_tool as oracle_tool

CURRENT_DB = acct_oracle.PROD_US
#SCHEMA = 'LIVE_CO'
SEED_FILE = '.\seed\Domain Data Template.xlsx'
#SEED_FILE = 'DDT.xlsx'

writer = pd.ExcelWriter('est.xlsx')

os.system("")
nameTime = time.strftime('%Y%m%d_%H%M%S')
excelName = 'est' + nameTime + '.xlsx'

workbook_ = load_workbook(SEED_FILE)

sheetnames =workbook_.get_sheet_names() 

query_sheet = workbook_.get_sheet_by_name('Checking_Query')

with UseOracleDB(CURRENT_DB) as cursor: