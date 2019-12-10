###
# Genearte data dictionary based on metadata and sample value
###

from openpyxl import Workbook
from openpyxl import load_workbook


from db_connect.sqlserver_db import UseSqlserverDB
import conf.acct as acct
import tool.TSQL as TSQL_function
import tool.tool as tool
import tool.oracle_tool as DB
from db_connect.oracle_db import UseOracleDB

import conf.acct_oracle as acct_oracle

CURRENT_DB = acct.UAT_UT_CAMPING_MART
SEED_FILE = r'.\seed\report.xlsx'
excelName = tool.file_name('DataDictionary','xlsx')
workbook = load_workbook(SEED_FILE)

CURRENT_DB = acct_oracle.QA_CDC


sheet = workbook.get_sheet_by_name('test')
rows = sheet.rows
columns = sheet.columns


query = """
with 
{schema}.sales as (
    SELECT to_char(ORD.ORD_DATE - 2/24,'mm') AS month_nb, to_char(ORD.ORD_DATE - 2/24,'Mon') AS month_dsc, LCLASS.LOCATION_CLASS_NAME, PRD.PRD_CD, PRD.PRD_NAME, ITEM.ID
    FROM {schema}.O_ORD_ITEM ITEM
    INNER JOIN {schema}.O_ORDER ORD ON ITEM.ORD_ID = ORD.ID 
    INNER JOIN {schema}.D_STORE DS ON ITEM.STORE_ID = DS.ID
    INNER JOIN {schema}.D_LOCATION_CLASS LCLASS ON LCLASS.ID = DS.LOCATION_CLASS_ID
    INNER JOIN {schema}.P_PRD PRD ON PRD.PRD_ID = ITEM.PRD_ID AND PRD.PRD_CD IN (
    'FP','FR'
    )
    WHERE ORD.ORD_DATE - 2/24 > '2019-01-01' AND ITEM.STATUS_ID = 1 -- Active
)

SELECT DISTINCT
s.month_dsc
,s.month_nb
,s.PRD_CD
,s.PRD_NAME
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'APOS' GROUP BY sc.month_dsc, sc.PRD_CD) AS APOS_CNT
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'Call Centre' GROUP BY sc.month_dsc, sc.PRD_CD) AS CC_CNT
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'District Office' GROUP BY sc.month_dsc, sc.PRD_CD) AS DO_CNT
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'Internet' GROUP BY sc.month_dsc, sc.PRD_CD) AS INT_CNT
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'Issuer' GROUP BY sc.month_dsc, sc.PRD_CD) AS ISSUER_CNT
,(SELECT COUNT(sc.ID) FROM sales sc WHERE sc.month_dsc = s.month_dsc AND sc.PRD_CD = s.PRD_CD AND sc.LOCATION_CLASS_NAME = 'Ministry Office' GROUP BY sc.month_dsc, sc.PRD_CD) AS MO_CNT


FROM {schema}.sales s
ORDER BY s.PRD_CD, s.month_nb
""".replace('{schema}','Live_AB_QA3')

print(query)

test = DB.query_db_pandas(query, CURRENT_DB)

print ("test")




'''for row in range(2,sheet.max_row+1):
    tableName = str(sheet.cell(row=row,column=1).value)
    columnName = str(sheet.cell(row=row,column=4).value)

    #print (tableName + ":" + columnName)
    query = "SELECT TOP 1 [" + columnName + "] FROM " + tableName + " WITH(NOLOCK) WHERE [" + columnName + "] IS NOT NULL"
    result = str(TSQL_function.inquery_single_row(query,cursor))
    sheet.cell(row=row,column=12).value = result
'''

workbook.save(excelName)  

