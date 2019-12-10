import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile

from openpyxl import Workbook
from openpyxl import load_workbook

df = pd.read_excel(r'.\seed\ab_seed.xlsx')


SEED_FILE = r'.\seed\report.xlsx'
workbook = load_workbook(SEED_FILE)

sheet = workbook.get_sheet_by_name('Tag Licenses')
rows = sheet.rows
columns = sheet.columns


for row in range(4,sheet.max_row+1):
        tableName = str(sheet.cell(row=row,column=1).value)
        columnName = str(sheet.cell(row=row,column=4).value)

        #print (tableName + ":" + columnName)

        sheet.cell(row=row,column=12).value = result

workbook.save(excelName)  