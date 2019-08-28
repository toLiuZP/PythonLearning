###
#  monitor Focus source
#  determine if the changes impact Mart
# TODO: add SVN version check.
###

import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
sys.path.append(os.getcwd())

import conf.acct_focus as acct
import db_connect.db_operator as DB
from db_connect.sqlserver_db import UseSqlserverDB, query_first_value, has_data, query, execute
from tool.tool import logger
from tool.send_mail import send_mail_outlook_html as mail
from tool.TSQL import query_meta_data
pd.set_option('max_colwidth',-1)

SEED_FILE = r"D:\Work\02. SVN\Aspira\SQL\AF_Data_Mart\trunk\DataModels\NJ_HF_MART.xlsx"
seed_workbook = load_workbook(SEED_FILE)

BASE_FILE = r".\maintain\monitor\Focus_Source_Base.xlsx"

LOG_FILE = r".\maintain\monitor\Focus_Change_Log.xlsx"
log_workbook = load_workbook(LOG_FILE)

TARGET_DB = acct.UAT_NJSTAGEUAT
meta = pd.DataFrame(columns = ['table_type','table_name','sql'])
ddl_pd = pd.DataFrame(columns =['ref_table','ref_column','typename','precision','scale','max_length','nullable','impact_table'])

@logger
def read_mapping(workbook):
    global meta
    sheet_list = ['Dimension','Bridge','Reference','Fact']
    for sheet_nm in sheet_list:
        sheet = workbook.get_sheet_by_name(sheet_nm)
        cell_nm = 'TABLE'
        for cell in sheet['C']:
            if cell.value != cell_nm:
                cell_nm = cell.value
                sql = sheet.cell(column=17,row = cell.row).value
                table_type = sheet.cell(column=1,row = cell.row).value
                meta = meta.append(pd.DataFrame({'table_type':[table_type],'table_name':[cell_nm],'sql':[sql]}),ignore_index=True)


def remove_first_comma(x): 
    if x.name == "impact_list":
        return [value[1:] for value in x.values]
    else:
        return x


@logger
def create_base(meta):
    global ddl_pd

    with UseSqlserverDB(TARGET_DB) as cursor:

        for index, row in meta.iterrows():
            tb_name = str(row['table_name'])
            sql = str(row['sql'])
            create_sql = "CREATE VIEW V_CDC_TEMP_"+ tb_name +" AS " + sql
            execute(cursor,create_sql)

            check_sql = "SELECT lower(a.referenced_entity_name) as ref_table ,lower(a.referenced_minor_name) as ref_column ,c.name as typename ,CONVERT(VARCHAR(50),b.precision) precision ,CONVERT(VARCHAR(50),b.scale) scale ,CONVERT(VARCHAR(50),b.max_length) max_length ,b.is_nullable nullable, '," + tb_name + "' AS impact_table FROM sys.dm_sql_referenced_entities ( 'DBO.V_CDC_TEMP_" + tb_name + "', 'OBJECT') a inner join sys.all_columns b on a.referenced_minor_name = b.name and a.referenced_id= b.object_id inner join sys.systypes c on b.system_type_id = c.xtype where a.referenced_minor_name is not null order by 1,2"
            
            signle = pd.DataFrame(query(cursor,check_sql))
            signle.columns = ['ref_table','ref_column','typename','precision','scale','max_length','nullable','impact_table']
            ddl_pd = ddl_pd.append(signle)

            drop_sql = "DROP VIEW V_CDC_TEMP_"+ tb_name
            execute(cursor,drop_sql)

    ddl_pd = ddl_pd.groupby(['ref_table','ref_column','typename','precision','scale','max_length','nullable']).agg(
        impact_list = pd.NamedAgg(column = 'impact_table', aggfunc = 'sum')
    )

    ddl_pd = ddl_pd.apply(remove_first_comma)
    ddl_pd.to_excel(BASE_FILE,sheet_name = "DDL")


@logger
def validate_base(log_wb,LOG_FILE):

    change_pd = pd.DataFrame(columns = ['change_date','change_type','impact_list','source_table_name','source_column_name','previous_column_type','new_column_type','previous_precision','new_precision','previous_scale','new_scale','previous_length','new_length','previous_nullable','new_nullable'])
    log_sheet = log_wb.get_sheet_by_name('Log')

    base_ddl = pd.read_excel(BASE_FILE)
    base_ddl['ref_table'] =base_ddl['ref_table'].fillna(method='ffill')

    table_list = ''
    for table_name in base_ddl['ref_table']:
        table_list = table_list + ",'" + table_name + "'"

    rs = pd.DataFrame(query_meta_data(table_list,TARGET_DB))
    rs.columns = ['ref_table','ref_column','typename','precision','scale','max_length','nullable']

    gap = pd.merge(base_ddl, rs, on = ['ref_table','ref_column'], how='outer')

    for index, col in gap.iterrows():
        if col[2] == col[8] and str(int(col[3])) == col[9] \
            and str(int(col[4])) == col[10] and str(int(col[5])) == col[11] and col[6] == col[12]:
            gap = gap.drop(index)

    gap = gap.sort_values(by = ['ref_table','ref_column'])

    for index, col in gap.iterrows():

        if str(col[8]) == 'nan':
            change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'change_type':['Deleted'],'impact_list':[col[7]],'source_table_name':[col[0]],'source_column_name':[col[1]],'previous_column_type':[col[2]],'new_column_type':[''],'previous_precision':[col[3]],'new_precision':[''],'previous_scale':[col[4]],'new_scale':[''],'previous_length':[col[5]],'new_length':[''],'previous_nullable':[col[6]],'new_nullable':['']}),ignore_index=True)
            log_sheet.append([datetime.date.today(),'Deleted',col[7],col[0],col[1],col[2],'',col[3],'',col[4],'',col[5],'',col[6],''])
        elif str(col[2]) != 'nan':
            change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'change_type':['Updated'],'impact_list':[col[7]],'source_table_name':[col[0]],'source_column_name':[col[1]],'previous_column_type':[col[2]],'new_column_type':[col[8]],'previous_precision':[col[3]],'new_precision':[col[9]],'previous_scale':[col[4]],'new_scale':[col[10]],'previous_length':[col[5]],'new_length':[col[11]],'previous_nullable':[col[6]],'new_nullable':[col[12]]}),ignore_index=True)
            log_sheet.append([datetime.date.today(),'Updated',col[7],col[0],col[1],col[2],col[8],col[3],col[9],col[4],col[10],col[5],col[11],col[6],col[12]])

    log_wb.save(LOG_FILE)
    return change_pd


if __name__ == '__main__':
    
    change_pd = validate_base(log_workbook,LOG_FILE)

    if not change_pd.empty:
        body = """
        Hi team,<br><br>
            Here is the NJ Source change list for today, please take a look.<br><br><br>
        <html><body>""" + change_pd.to_html(index=False) + '</body></html>' 
        attachments = [os.getcwd()+LOG_FILE[1:]]
        mail('(Auto Generation) NJ Source Change List',['zongpei.liu@aspiraconnect.com;zongpei.liu@aspiraconnect.com;Tom.Xie@aspiraconnect.com;Gary.Zhou@aspiraconnect.com;Tim.Wang@aspiraconnect.com;Kelvin.Wang@aspiraconnect.com'],body,attachments)
        #mail('(Auto Generation) NJ Source Change List',['zongpei.liu@aspiraconnect.com'],body,attachments)
        read_mapping(seed_workbook)
        create_base(meta)
    else:
        body = """
        Hi team,<br><br>
            Everything is good for NJ Source Change List.
        <html>
        <body> </body></html>"""
        mail('(Auto Generation) All good for NJ Source Change List',['zongpei.liu@aspiraconnect.com'],body)
    