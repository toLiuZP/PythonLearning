###
#  monitor Focus source
#  determine if the changes impact Mart
# 
# 
# 
# 
###

import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
import openpyxl.styles as sty
import os
import sys
sys.path.append(os.getcwd())

import conf.acct_focus as acct
import db_connect.db_operator as DB
from db_connect.sqlserver_db import UseSqlserverDB, query_first_value, has_data, query, execute
from tool.tool import file_name,logger,identify_backup_tables
from tool.send_mail import send_mail_outlook_html as mail

SEED_FILE = r"D:\Work\02. SVN\Aspira\SQL\AF_Data_Mart\trunk\DataModels\NJ_HF_MART.xlsx"
seed_workbook = load_workbook(SEED_FILE)

BASE_FILE = r".\seed\Focus_Source_Base.xlsx"
excelName = r".\output\Focus_Source.xlsx"
base_workbook = load_workbook(BASE_FILE)

LOG_FILE = r".\seed\Focus_Change_Log.xlsx"
log_workbook = load_workbook(LOG_FILE)

TARGET_DB = acct.QA_NJDMAQA
meta = pd.DataFrame(columns = ['table_type','table_name','sql'])


def read_mapping(workbook):
    global meta
    sheet_list = ['Dimension','Bridge','Reference','Fact']
    for sheet_nm in sheet_list:
        sheet = workbook.get_sheet_by_name(sheet_nm)
        cell_nm = 'TABLE'
        for cell in sheet['C']:
            if cell.value != cell_nm:
                cell_nm = cell.value
                sql = sheet.cell(column=17,row = cell.row).value
                table_type = sheet.cell(column=1,row = cell.row).value
                meta = meta.append(pd.DataFrame({'table_type':[table_type],'table_name':[cell_nm],'sql':[sql]}),ignore_index=True)

@logger
def create_base(workbook,meta):

    all_sheet = workbook.create_sheet(title = 'ALL_DDL')
    all_pd = pd.DataFrame(columns = ['ref_table','ref_column','typename','precision','scale','max_length','nullable','impact_list'])

    with UseSqlserverDB(TARGET_DB) as cursor:

        for index, row in meta.iterrows():
            tb_name = str(row['table_name'])
            sql = str(row['sql'])
            create_sql = "CREATE VIEW V_CDC_TEMP_"+ tb_name +" AS " + sql
            execute(cursor,create_sql)

            check_sql = "SELECT a.referenced_entity_name as ref_table ,a.referenced_minor_name as ref_column ,c.name as typename ,CONVERT(VARCHAR(50),b.precision) precision ,CONVERT(VARCHAR(50),b.scale) scale ,CONVERT(VARCHAR(50),b.max_length) max_length ,b.is_nullable nullable, '" + tb_name + "' AS impact_list FROM sys.dm_sql_referenced_entities ( 'DBO.V_CDC_TEMP_" + tb_name + "', 'OBJECT') a inner join sys.all_columns b on a.referenced_minor_name = b.name and a.referenced_id= b.object_id inner join sys.systypes c on b.system_type_id = c.xtype where a.referenced_minor_name is not null order by 1,2"
            rs = query(cursor,check_sql)

            drop_sql = "DROP VIEW V_CDC_TEMP_"+ tb_name
            execute(cursor,drop_sql)

            sheet = workbook.create_sheet(title = tb_name[:30])
            sheet.append(['ref_table','ref_column','typename','precision','scale','max_length','nullable'])

            for row in rs:
                ref_table = str(row[0])
                ref_column = str(row[1])
                typename = str(row[2])
                precision = str(row[3])
                scale = str(row[4])
                max_length = str(row[5])
                nullable = str(row[6])
                impact_table = str(row[7])
                sheet.append([ref_table,ref_column,typename,precision,scale,max_length,nullable])
                all_pd = all_pd.append(pd.DataFrame({'ref_table':[ref_table],'ref_column':[ref_column],'typename':[typename],'precision':[precision],'scale':[scale],'max_length':[max_length],'nullable':[nullable],'impact_list':[impact_table]}),ignore_index=True)
    
    all_pd = all_pd.sort_values(by=['ref_table','ref_column'])
    all_pd = all_pd.reset_index(drop=True)
    last_table_name = ''
    last_column_name = ''
    table_list = ''

    for index, row in all_pd.iterrows():
        if not last_table_name:
            last_table_name = row['ref_table']
            last_column_name = row['ref_column']
            last_type_name = row['typename']
            last_precision = row['precision']
            last_scale = row['scale']
            last_length = row['max_length']
            last_nullable = row['nullable']
            last_impact_table_name = row['impact_list']
            table_list = last_impact_table_name
        else:
            if last_table_name != row['ref_table']:
                all_sheet.append([last_table_name,last_column_name,last_type_name,last_precision,last_scale,last_length,last_nullable,table_list])
                table_list = ''
                last_table_name = row['ref_table']
                last_column_name = row['ref_column']
                last_type_name = row['typename']
                last_precision = row['precision']
                last_scale = row['scale']
                last_length = row['max_length']
                last_nullable = row['nullable']
                last_impact_table_name = row['impact_list']
                table_list = last_impact_table_name
            else:
                if last_column_name == row['ref_column']:
                    table_list = table_list + "," + row['impact_list']
                else:
                    all_sheet.append([last_table_name,last_column_name,last_type_name,last_precision,last_scale,last_length,last_nullable,table_list])
                    table_list = ''
                    table_list = last_impact_table_name
                    last_table_name = row['ref_table']
                    last_column_name = row['ref_column']
                    last_type_name = row['typename']
                    last_precision = row['precision']
                    last_scale = row['scale']
                    last_length = row['max_length']
                    last_nullable = row['nullable']
                    last_impact_table_name = row['impact_list']


@logger
def validate_base(workbook,log_wb,LOG_FILE):

    change_pd = pd.DataFrame(columns = ['change_date','impact_list','source_table_name','source_column_name','previous_column_type','new_column_type','previous_precision','new_precision','previous_scale','new_scale','previous_length','new_length','previous_nullable','new_nullable'])
    log_sheet = log_wb.get_sheet_by_name('Log')
    all_sheet = workbook.get_sheet_by_name('ALL_DDL')
    table_list = ''
    rs =[]
    for cell in all_sheet['A']:
        table_list = table_list + ",'" + cell.value + "'"

    sql = "SELECT a.name as ref_table ,b.name as ref_column ,c.name as typename ,CONVERT(VARCHAR(50),b.precision) precision ,CONVERT(VARCHAR(50),b.scale) scale ,CONVERT(VARCHAR(50),b.max_length) max_length ,b.is_nullable nullable FROM sys.all_objects a inner join sys.all_columns b on a.object_id= b.object_id inner join sys.systypes c on b.system_type_id = c.xtype WHERE a.name in (" + table_list[1:] +") ORDER BY 1,2"

    with UseSqlserverDB(TARGET_DB) as cursor:
        rs = query(cursor,sql)

    for row in all_sheet.rows:
        for line in rs:
            if row[0].value == line[0] and row[1].value == line[1]:
                if row[2].value != line[2] or row[3].value != line[3] or row[4].value != line[4] or row[5].value != line[5] or row[6].value != str(line[6]):
                    change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'impact_list':[row[7].value],'source_table_name':[row[0].value],'source_column_name':[row[1].value],'previous_column_type':[row[2].value],'new_column_type':[line[2]],'previous_precision':[row[3].value],'new_precision':[line[3]],'previous_scale':[row[4].value],'new_scale':[line[4]],'previous_length':[row[5].value],'new_length':[line[5]],'previous_nullable':[row[6].value],'new_nullable':[line[6]]}),ignore_index=True)
                    log_sheet.append([datetime.date.today(),row[7].value,row[0].value,row[1].value,row[2].value,line[2],row[3].value,line[3],row[4].value,line[4],str(row[5].value),str(line[5]),row[6].value.upper(),line[6]])
                break
        
    log_wb.save(LOG_FILE)
    return change_pd

def highlight_last_row(s):
    return ['background-color: #FF0000' if i==len(s)-1 else '' for i in range(len(s))]

if __name__ == '__main__':

    #read_mapping(seed_workbook)
    #create_base(base_workbook,meta)
    
    #base_workbook.save(excelName)

    change_pd = validate_base(base_workbook,log_workbook,LOG_FILE)
    pd.set_option('max_colwidth',-1)

    if not change_pd.empty:
        body = """
        Hi team,<br><br>
            Here is the NJ Source change list for today, please take a look.<br><br><br>
        
        <html>
            <meta  charset=utf-8"/>
                <style>
                    mark {
                        background-color:#00ff90; font-weight:bold;
                    }
                </style>
        <body>""" + change_pd.to_html(index_names=False) + '</body></html>' 
        #<body>""" + change_pd.style.apply(highlight_last_row).to_html(index_names=False) + '</body></html>' 
        attachments = [os.getcwd()+LOG_FILE[1:]]
        #mail('NJ Source Change List',['zongpei.liu@aspiraconnect.com;zongpei.liu@aspiraconnect.com;Tom.Xie@aspiraconnect.com;Gary.Zhou@aspiraconnect.com;Tim.Wang@aspiraconnect.com;Kelvin.Wang@aspiraconnect.com'],body,attachments)
        mail('NJ Source Change List',['zongpei.liu@aspiraconnect.com'],body,attachments)
