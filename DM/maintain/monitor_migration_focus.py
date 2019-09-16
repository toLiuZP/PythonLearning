###
#  monitor Focus Migration target table
#  determine if the changes impact Mart
#  TODO: add re_name and drop monitor.
###

import pandas as pd
import numpy as np 
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
sys.path.append(os.getcwd())

import conf.acct_focus as acct
from db_connect.sqlserver_db import UseSqlserverDB, query_first_value, has_data, query, execute
from tool.tool import logger
from tool.send_mail import send_mail_outlook_html as mail
from tool.TSQL import query_meta_data
pd.set_option('max_colwidth',-1)

MINITOR_TABLE_LIST = [
'ActiveOutdoorsTransaction',
'Address2',
'AddressDetail',
'Agent',
'AgentApplication',
'AgentBankAccount',
'AgentOutletApplicationOutlet',
'AgentOwnership',
'ChildSupport',
'ChildSupportAnswer',
'Customer',
'CustomerBusiness',
'CustomerBusinessOwnership',
'CustomerDisability',
'CustomerHunterEd',
'CustomerIdentity',
'CustomerIndividual',
'CustomerQualification',
'CustomerQualificationFarmerCertification',
'CustomerRestriction',
'CustomerRestrictionDetail',
'CustomerVessel',
'CustomerVesselGearType',
'CustomerVesselOwnership',
'Document',
'DocumentDetail',
'DrawTicket',
'DrawTicketHuntChoice',
'FeeGroup',
'Hunt',
'HuntApplication',
'HuntApplicationLicense',
'HuntSeason',
'HuntSeasonHunt',
'HuntTypeLicenseYear',
'Item',
'ItemAmountCalculateUnit',
'ItemFee',
'ItemFeeDistribution',
'ItemPackage',
'ItemQuestion',
'ItemQuestionActionableAnswer',
'ItemRule',
'ItemRuleParameter',
'ItemSalesBlackout',
'LEPermit',
'License',
'LicenseAction',
'LicenseAncillaryData',
'LicenseReport',
'LicenseReportAnswer',
'LicenseReportAnswerGroup',
'LicenseReportAnswerRow',
'LicenseUnitEntry',
'Outlet',
'OutletBusinessHours',
'OutletPaymentMethod',
'OutletTypeItem',
'Person',
'Phone',
'PrintTemplate',
'RootItemNumber',
'SchedulePattern',
'Transaction',
'TransactionDetail',
'TransactionDetailDistribution',
'TransactionDetailFee',
'TransactionHeader'
]

BASE_FILE = r".\maintain\monitor\Focus_Migration_Source_Base.xlsx"
base_workbook = load_workbook(BASE_FILE)

LOG_FILE = r".\maintain\monitor\Focus_Migration_Change_Log.xlsx"
log_workbook = load_workbook(LOG_FILE)

TARGET_DB = acct.QA_NJMAINQA


def create_base():
    global MINITOR_TABLE_LIST
    
    table_list = ''
    for table_name in MINITOR_TABLE_LIST:
        table_list += ",'" + table_name + "'"

    df = pd.DataFrame(query_meta_data(table_list,TARGET_DB))
    df.columns = ['ref_table','ref_column','typename','precision','scale','max_length','nullable']
    df.to_excel(BASE_FILE,sheet_name = "DDL",index=False)


@logger
def validate_base(workbook,log_wb,LOG_FILE):

    change_pd = pd.DataFrame(columns = ['change_date','change_type','source_table_name','source_column_name','previous_column_type','new_column_type','previous_precision','new_precision','previous_scale','new_scale','previous_length','new_length','previous_nullable','new_nullable'])
    log_sheet = log_wb.get_sheet_by_name('Log')
    all_sheet = pd.read_excel(BASE_FILE)
    table_list = ''
    for cell in all_sheet['ref_table']:
        table_list = table_list + ",'" + cell + "'"
 
    rs = pd.DataFrame(query_meta_data(table_list,TARGET_DB))
    rs.columns = ['ref_table','ref_column','typename','precision','scale','max_length','nullable']

    gap = pd.merge(all_sheet, rs, on = ['ref_table','ref_column'], how='outer')

    for index, col in gap.iterrows():
        if col[2] == col[7] and str(int(col[3])) == col[8] \
            and str(int(col[4])) == col[9] and str(int(col[5])) == col[10] and col[6] == col[11]:
            gap = gap.drop(index)

    gap = gap.sort_values(by = ['ref_table','ref_column'])

    for index, col in gap.iterrows():

        if str(col[2]) == 'nan':
            change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'change_type':['Added'],'source_table_name':[col[0]],'source_column_name':[col[1]],'previous_column_type':[' '],'new_column_type':[col[7]],'previous_precision':[' '],'new_precision':[col[8]],'previous_scale':[' '],'new_scale':[col[9]],'previous_length':[' '],'new_length':[col[10]],'previous_nullable':[' '],'new_nullable':[col[11]]}),ignore_index=True)
            log_sheet.append([datetime.date.today(),'Added',col[0],col[1],'',col[7],'',col[8],'',col[9],'',col[10],'',col[11]])
        elif str(col[7]) == 'nan':
            change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'change_type':['Deleted'],'source_table_name':[col[0]],'source_column_name':[col[1]],'previous_column_type':[col[2]],'new_column_type':[''],'previous_precision':[col[3]],'new_precision':[''],'previous_scale':[col[4]],'new_scale':[''],'previous_length':[col[5]],'new_length':[''],'previous_nullable':[col[6]],'new_nullable':['']}),ignore_index=True)
            log_sheet.append([datetime.date.today(),'Deleted',col[0],col[1],col[2],'',col[3],'',col[4],'',col[5],'',col[6],''])
        else:
            change_pd = change_pd.append(pd.DataFrame({'change_date':[datetime.date.today()],'change_type':['Updated'],'source_table_name':[col[0]],'source_column_name':[col[1]],'previous_column_type':[col[2]],'new_column_type':[col[7]],'previous_precision':[col[3]],'new_precision':[col[8]],'previous_scale':[col[4]],'new_scale':[col[9]],'previous_length':[col[5]],'new_length':[col[10]],'previous_nullable':[col[6]],'new_nullable':[col[11]]}),ignore_index=True)
            log_sheet.append([datetime.date.today(),'Updated',col[0],col[1],col[2],col[7],col[3],col[8],col[4],col[9],col[5],col[10],col[6],col[11]])
      
    log_wb.save(LOG_FILE)
    return change_pd


if __name__ == '__main__':

    change_pd = validate_base(base_workbook,log_workbook,LOG_FILE)
    if not change_pd.empty:
        body = """
        Hi team,<br><br>
            Here is the NJ Migration Tables Change List for today, please take a look.<br><br><br>
        <html>
        <body>""" + change_pd.to_html(index=False) + '</body></html>' 
        attachments = [os.getcwd()+LOG_FILE[1:]]
        mail('(Auto Generation) NJ Migration Tables Change List',['zongpei.liu@aspiraconnect.com;Tom.Xie@aspiraconnect.com;Aspira_DMA_AspiraFocus_Migration@aspiraconnect.com'],body,attachments)
        #mail('(Auto Generation) NJ Migration Tables Change List',['zongpei.liu@aspiraconnect.com'],body,attachments)
        create_base()
    else :
        body = """
        Hi team,<br><br>
            Everything is good for NJ Migration Tables List.
        <html>
        <body> </body></html>"""
        mail('(Auto Generation) All good for NJ Migration Tables List',['zongpei.liu@aspiraconnect.com'],body)
        
    