import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
import openpyxl.styles as sty

import conf.acct_oracle as acct
from db_connect.oracle_db import UseOracleDB
import tool.oracle_tool as oracle_tool
from tool.tool import file_name,logger 

SEED_FILE = ".\seed\ORACLE_DDL_GAP.xlsx"
nameTime = time.strftime('%Y%m%d_%H%M%S')
excelName = file_name('ORACLE_DDL_GAP','xlsx')
workbook = load_workbook(SEED_FILE)
ddl_sheet = workbook.get_sheet_by_name('DDL')
synonyms_sheet = workbook.get_sheet_by_name('SYNONYMS')


def merge(dev, qa, sheet, key_column):

    dev = dev.fillna('null')
    qa = qa.fillna('null')
    gap = pd.merge(dev, qa, on = key_column, how='outer')
    
    for index, col in gap.iterrows():
        if str(col[1]) == 'nan' or str(col[2]) == 'nan':
            pass
        else:
            gap = gap.drop(index)

    gap = gap.sort_values(by = key_column)
    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+2,column=i+1).value = col[i]
            if i == 1:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i == 2:   
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")

def check_ddl(workbook,sheet,db_a,schema_a,name_a,db_b, schema_b,name_b):

    query_a = "SELECT TABLE_NAME, '" + name_a + "' AS OWNER FROM ALL_TAB_COMMENTS WHERE OWNER = '" + schema_a + "' AND TABLE_TYPE = 'TABLE' AND TABLE_NAME NOT LIKE 'BIN$%' ORDER BY TABLE_NAME"
    query_b = "SELECT TABLE_NAME, '" + name_b + "' AS OWNER FROM ALL_TAB_COMMENTS WHERE OWNER = '" + schema_b + "' AND TABLE_TYPE = 'TABLE' AND TABLE_NAME NOT LIKE 'BIN$%' ORDER BY TABLE_NAME"

    ddl_a = oracle_tool.query_db_pandas(query_a, db_a)
    ddl_b = oracle_tool.query_db_pandas(query_b, db_b)

    merge(ddl_a,ddl_b,sheet,['TABLE_NAME'])

def check_synonyms(workbook,sheet,db_a,schema_a,name_a,db_b,schema_b,name_b):

    query_a = "SELECT SYNONYM_NAME, '" + name_a + "' AS OWNER FROM all_synonyms WHERE OWNER = '" + schema_a + "' ORDER BY SYNONYM_NAME"
    query_b = "SELECT SYNONYM_NAME, '" + name_b + "' AS OWNER FROM all_synonyms WHERE OWNER = '" + schema_b + "' ORDER BY SYNONYM_NAME"

    synonyms_a = oracle_tool.query_db_pandas(query_a, db_a)
    synonyms_b = oracle_tool.query_db_pandas(query_b, db_b)

    merge(synonyms_a,synonyms_b,sheet,['SYNONYM_NAME'])

if __name__ == '__main__':

    db_a = acct.QA_CDC
    name_a = 'CO'
    schema_a = 'LIVE_CO_QA3'
    db_b = acct.QA_CDC
    name_b = 'TX'
    schema_b = 'LIVE_TX_QA3'

    check_ddl(workbook,ddl_sheet,db_a,schema_a,name_a,db_b,schema_b,name_b)
    check_synonyms(workbook,synonyms_sheet,db_a,schema_a,name_a,db_b,schema_b,name_b)
    
workbook.save(excelName)
