###
# 1. Search all database name in one server
# 2. Search all columns named STREET_ADR_LN_1 and its type is varchar(512) 
# 2.1 generate scripts to alter these columns to varchar(3000)
# 3. Search all columns named CITY_NM and its type is varchar(255)
# 3.1 generate scripts to alter these columns to varchar(512)
# 
###

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
import openpyxl.styles as sty

import conf.acct as acct
import db_connect.db_operator as DB
from db_connect.sqlserver_db import UseSqlserverDB
from tool.tool import file_name,logger, save_file
from tool.TSQL import search_db, search_column

SEED_FILE = ".\seed\Search_columns.xlsx"
nameTime = time.strftime('%Y%m%d_%H%M%S')
excelName = file_name('Search_Columns','xlsx')
workbook = load_workbook(SEED_FILE)
tempate_sheet = workbook.get_sheet_by_name('Template')

SERVER_LIST = (
    #(acct.DEV_CO_HF_MART, 'DEV'),
    #(acct.QA_CO_HF_MART, 'QA_US'),
    #(acct.QA_AB_HF_MART, 'QA_CA'),
    #(acct.UAT_CO_HF_MART, 'UAT_US'),
    #(acct.UAT_AB_HF_MART, 'UAT_CA'),
    (acct.PROD_CO_HF_MART, 'PROD_US'),
    (acct.PROD_AB_HF_MART, 'PROD_CA')
)


def gen_alter_sql(table_nm, column_nm, type_nm, length, new_type, new_len)->str:
    sql = '''
IF EXISTS(select * from sysobjects a,syscolumns b,systypes c where a.id=b.id and a.name='table_nm' and a.xtype='U' and b.name = 'column_nm' and b.xtype=c.xtype and b.length = [length] and c.name = 'type_nm')
BEGIN
	ALTER TABLE DBO.table_nm ALTER COLUMN column_nm new_type(new_len) null
	PRINT '[INFO] ALTER COLUMN [DBO].[table_nm].[column_nm] TO new_type(new_len)'
END

'''.replace('table_nm',table_nm).replace('column_nm',column_nm).replace('type_nm',type_nm).replace('[length]',length).replace('new_type',new_type).replace('new_len',new_len)
    return sql



if __name__ == '__main__':

    for server in SERVER_LIST:
        server_nm = server[0]
        name = server[1]

        with UseSqlserverDB(server_nm) as cursor:
            db_list = search_db(cursor)

            gen_sql = ''
            rollback_sql = ''
            count = 1

            new_sheet = workbook.copy_worksheet(tempate_sheet)
            new_sheet.title = name

            for db in db_list:
                address1_list = search_column(cursor, db_name = db[0], column_name = "a.name LIKE 'STREET_ADR_LN_1' OR a.name LIKE 'ADDRESS_LINE_1%' ", column_type = 'varchar',column_len = '512', table_name = 'D_ADDRESS')
                city_list = search_column(cursor, db_name = db[0], column_name = "a.name = 'CITY_NM' OR a.name = 'CITY_NAME' ", column_type = 'varchar',column_len = '255', table_name = 'D_ADDRESS')

                if len(address1_list) > 0 or len(city_list) > 0:            
                    gen_sql += 'USE ' + db[0] + '\nGO \n'
                    rollback_sql += 'USE ' + db[0] + '\nGO \n'

                    for row in address1_list:
                        count += 1

                        gen_sql += gen_alter_sql(table_nm = str(row[1]), column_nm = str(row[3]) , type_nm = str(row[6]), length = str(row[7]), new_type = 'varchar', new_len = '3000')
                        rollback_sql += gen_alter_sql(table_nm = str(row[1]), column_nm = str(row[3]) , type_nm = 'varchar', length = '3000', new_type = str(row[6]), new_len = str(row[7]))

                        for i in range(1, len(row), 1):
                            if i == 1:
                                new_sheet.cell(row=count,column=i).value = db[0]
                            else:
                                new_sheet.cell(row=count,column=i).value = row[i-1]

                    for row in city_list:
                        count += 1

                        gen_sql += gen_alter_sql(table_nm = str(row[1]), column_nm = str(row[3]) , type_nm = str(row[6]), length = str(row[7]), new_type = 'varchar', new_len = '512')
                        rollback_sql += gen_alter_sql(table_nm = str(row[1]), column_nm = str(row[3]) , type_nm = 'varchar', length = '512', new_type = str(row[6]), new_len = str(row[7]))

                        for i in range(1, len(row), 1):
                            if i == 1:
                                new_sheet.cell(row=count,column=i).value = db[0]
                            else:
                                new_sheet.cell(row=count,column=i).value = row[i-1]

            #save_file(file_name("Extend_Address_"+name,".sql"),gen_sql)
            #save_file(file_name("Rollback_Extend_Address_"+name,".sql"),rollback_sql)

workbook.remove_sheet(workbook.get_sheet_by_name('Template'))
workbook.save(excelName)
