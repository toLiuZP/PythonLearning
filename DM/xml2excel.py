###
# Convert Jasper XML to Excel file 
#         
###

from openpyxl import Workbook
xml_file = r"C:\test\nj_dev_transaction_domain.xml"
output = r"C:\test\domain_dictionary.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Domain Dictionary"

ws.cell(row=1,column=1).value = 'Entity Category'
ws.cell(row=1,column=2).value = 'Entity Name'
ws.cell(row=1,column=3).value = 'Attribute Name'
ws.cell(row=1,column=4).value = 'Definition'
ws.cell(row=1,column=5).value = 'Aspira Focus Location'
ws.cell(row=1,column=6).value = 'Source'
row_number = 2


def find_value(line,key_word):
    start_postion = line.find(key_word)
    if start_postion != -1:
        line = line[start_postion+len(key_word):]
        end_position = line.find('"')
        key_value = line[:end_position]
        return key_value


with open(xml_file,encoding="utf") as file_object:
    lines = file_object.readlines()
for line in lines:
    line = line.lstrip()

    if line.startswith('<itemGroup description="" descriptionId="" id="'):
        if line.startswith('<itemGroup description="" descriptionId="" id="newSet'):
            ws.cell(row=row_number,column=2).value = find_value(line,'label="')
        else:
            ws.cell(row=row_number,column=1).value = find_value(line,'label="')
        row_number +=1

    if line.startswith('<item defaultAgg="'):
        ws.cell(row=row_number,column=3).value = find_value(line,'label="')
        field_id = find_value(line,'resourceId="JoinTree_1.')
        id_value = find_value(line,'id="')
        for source in lines:
            source = source.lstrip()
            if source.startswith('<field id="'+field_id):
                data_set_expression = find_value(source,'dataSetExpression="')
                ws.cell(row=row_number,column=6).value = data_set_expression if data_set_expression else find_value(line,'resourceId="')   
        row_number +=1
    

wb.save(output)