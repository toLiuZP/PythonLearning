###
# generate DDL based on mapping (NJ version)
# TODO: convert column type from mssql to oracle
# TODO: print column length > 30
# TODO: handle PK for non-IDENTITY tables

# Zongpei: added spell checking function
###

import os

from openpyxl import Workbook
from openpyxl import load_workbook
import time
from spellchecker import SpellChecker

from tool.tool import save_file

os.system("")

SEED_FILE = r".\seed\Mapping.xlsx"
nameTime = time.strftime('%m/%d/%Y')
workbook = load_workbook(SEED_FILE)
sheetnames = workbook.get_sheet_names() 

gen_tables = ['D_AGE_CATEGORY','R_SALES_CHANNEL','R_ORDER_CATEGORY','D_CUSTOMER','D_PRODUCT','F_ACTIVITY_SALES','F_MONTHLY_SALES']
#gen_tables = ['B_AGENT_BANK_ACCOUNT','B_AGENT_CONTACT','B_AGENT_OWNER','B_CUSTOMER_BUSINESS_OWNERSHIP','B_CUSTOMER_CHILD_SUPPORT','B_CUSTOMER_HUNTER_EDUCATION','B_CUSTOMER_LEPERMIT','B_CUSTOMER_QUALIFICATION','B_CUSTOMER_RESTRICTION','B_GROUP_MEMBERSHIP','B_GROUP_PERMISSION','B_HUNT_APPLICATION_CHOICE','B_HUNT_APPLICATION_CUSTOMER','B_HUNT_AREA_COUNTY','B_HUNT_SEASON_HUNT','B_HUNT_TYPE_LICENSE_YEAR_HUNT','B_HUNT_TYPE_LICENSE_YEAR_HUNT_GENERATION','B_ITEM_PACKAGE','B_ITEM_PROPERTIES','B_ITEM_QUESTION','B_ITEM_QUESTION_ANSWER','B_LICENSE_ANCILLARY_DATA','B_LICENSE_REPORT_QUESTION_ACTION','B_LICENSE_REPORT_TEMPLATE_USAGE_LEVEL','B_USER_OUTLET','B_VESSEL_DOCUMENTATION','B_VESSEL_HOME_PORT','B_VESSEL_LEASE','B_VESSEL_OWNERSHIP','D_ADDRESS','D_AGENT','D_AGENT_APPLICATION_BUSINESS_INFO','D_ANSWER_OPTION','D_AUDIT_TRANSACTION_LOG','D_COUNTY','D_CUSTOMER','D_CUSTOMER_IDENTITY','D_DATE','D_DOCUMENT','D_DRAW','D_GROUP','D_HUNT','D_HUNT_APPLICATION','D_HUNT_AREA','D_HUNT_TYPE_LICENSE_YEAR','D_ITEM','D_LE_PERMIT','D_LE_PERMIT_TYPE','D_LICENSE','D_LICENSE_REPORT','D_LICENSE_REPORT_QUESTION','D_LICENSE_REPORT_QUESTION_GROUP','D_LICENSE_REPORT_TEMPLATE','D_OUTLET','D_PERMISSION','D_QUESTION','D_TIME','D_USER','D_VESSEL_PORT','F_AGENT_APPLICATION','F_HUNT_TYPE_LICENSE_YEAR_DRAW_STATISTICS','F_LICENSE_ACTION','F_PERMIT_CUSTOMER_TRANSFER','F_TRANSACTION_DETAIL','R_CUSTOMER_SOURCE','R_GENDER','R_LAND_TYPE','R_LICENSE_ACTION_TYPE','R_PERMIT_CUSTOMER_TRANSFER_TYPE','R_SALES_CHANNEL','R_STATUS_CODE','R_TRANSACTION_DETAIL_TYPE','R_TRANSACTION_TYPE','R_WEAPON','RPT_FEDERAL_AID','RPT_TRANSACTION_SALES']

HEADER = '''
/*
 * NOTES: Creates [RAPLACE_TABLE_NM] for AspiraFocus datamart 
 *
 * DATE      	JIRA    	USER       		DESCRIPTION
 * ----------	--------	-----------		---------------------------------------
 * [RAPLACE_TODAY]	DMA-4696	Zongpei Liu		Initialization.
*/

SET NOCOUNT ON

SET TRANSACTION ISOLATION LEVEL READ COMMITTED

IF OBJECT_ID('DBO.[RAPLACE_TABLE_NM]') IS NULL
BEGIN
	CREATE TABLE DBO.[RAPLACE_TABLE_NM](
'''

PK_IDENTITY_COLUMN = '        [RAPLACE_COLUMN_NM][RAPLACE_COLUMN_TYPE]IDENTITY(1,1),\n'
N_COLUMN = '        [RAPLACE_COLUMN_NM][RAPLACE_COLUMN_TYPE]NULL,\n'
COMMENT = "	exec sys.sp_addextendedproperty 'MS_Description', '[RAPLACE_COMMENT]', 'schema', 'dbo', 'table', '[RAPLACE_TABLE_NM]', 'column', '[RAPLACE_COLUMN_NM]'\n"
INDEX = '''
IF NOT EXISTS( SELECT TOP 1 1 FROM sys.indexes i WHERE i.object_id = object_id('[dbo].[[RAPLACE_TABLE_NM]]','U') AND i.name = '[RAPLACE_TABLE_NM]_[RAPLACE_COLUMN_NM]_IDX')
BEGIN
    CREATE NONCLUSTERED INDEX [[RAPLACE_TABLE_NM]_[RAPLACE_COLUMN_NM]_IDX] ON [dbo].[[RAPLACE_TABLE_NM]]([[RAPLACE_COLUMN_NM]])
    WITH (PAD_INDEX= OFF,STATISTICS_NORECOMPUTE =OFF, SORT_IN_TEMPDB = ON, DROP_EXISTING= OFF, MAXDOP=0, ALLOW_ROW_LOCKS = ON, ALLOW_PAGE_LOCKS= ON, DATA_COMPRESSION=PAGE)
    ON {INDEXFG}
    PRINT '[INFO] CREATED NONCLUSTERED INDEX [dbo].[[RAPLACE_TABLE_NM]].[[RAPLACE_TABLE_NM]_[RAPLACE_COLUMN_NM]_IDX]'
END
'''



ORACLE_HEADER = '''
-- TABLE: {SCHEMA}.[RAPLACE_TABLE_NM]
CREATE TABLE {SCHEMA}.[RAPLACE_TABLE_NM](
'''
ORACLE_COLUMN = '    [RAPLACE_COLUMN_NM][RAPLACE_COLUMN_TYPE],\n'
ORACLE_COMMENT = "COMMENT ON COLUMN {SCHEMA}.[RAPLACE_TABLE_NM].[RAPLACE_COLUMN_NM] IS '[RAPLACE_COMMENT]';\n"
ORACLE_PK_INDEX = '''
CREATE UNIQUE INDEX {SCHEMA}.PK_[RAPLACE_TABLE_NM] ON {SCHEMA}.[RAPLACE_TABLE_NM]([RAPLACE_PK_NM]) TABLESPACE {INDEXSPACE};\n
'''
ORACLE_PK_CONSTRAINT = '''
ALTER TABLE {SCHEMA}.[RAPLACE_TABLE_NM] ADD 
    CONSTRAINT PK_[RAPLACE_TABLE_NM] PRIMARY KEY ([RAPLACE_PK_NM])
    USING INDEX {SCHEMA}.PK_[RAPLACE_TABLE_NM] ;\n'''

def beauty_empty(statement, column_nm, column_type):
    for _ in range(1,50 - len(column_nm)):
        column_nm += ' '
    for _ in range(1,20 - len(column_type)):
        column_type += ' '
    return statement.replace('[RAPLACE_COLUMN_NM]',column_nm).replace('[RAPLACE_COLUMN_TYPE]',column_type)


def beauty_empty_oracle(statement, column_nm, column_type):
    for _ in range(1,35 - len(column_nm)):
        column_nm += ' '
    return statement.replace('[RAPLACE_COLUMN_NM]',column_nm).replace('[RAPLACE_COLUMN_TYPE]',column_type)


def generate_ms_ddl():
    for table_nm in gen_tables:

        check_name(table_nm, None)

        ddl = ''
        comment_text = ''
        row_count = 0
        pk_column = ''
        index = ''
        ddl += HEADER.replace('[RAPLACE_TABLE_NM]',table_nm).replace('[RAPLACE_TODAY]',nameTime) 

        for sheetname in sheetnames:
            table_found = False
            pk_found = False
            if sheetname != 'ChangeLog':
                sheet = workbook.get_sheet_by_name(sheetname)
                for row in range(1,sheet.max_row+1):
                    if sheet.cell(row=row,column=3).value == table_nm:
                        table_found = True
                        column_nm = str(sheet.cell(row=row,column=4).value)
                        column_type = str(sheet.cell(row=row,column=6).value)
                        if sheet.cell(row=row,column=5).value != None:
                            check_name(column_nm,column_type)
                        pk_ind = str(sheet.cell(row=row,column=7).value)
                        column_comment = str(sheet.cell(row=row,column=19).value)
                        index_ind = str(sheet.cell(row=row,column=20).value)
                        row_count +=1
                        if row_count == 2:
                            ddl += beauty_empty(PK_IDENTITY_COLUMN,column_nm,column_type)
                            if column_comment != 'None':
                                comment_text += COMMENT.replace('[RAPLACE_COLUMN_NM]',column_nm).replace('[RAPLACE_TABLE_NM]',table_nm).replace('[RAPLACE_COMMENT]',column_comment)
                            if pk_ind == 'PK':
                                pk_column += ',' + column_nm 
                                pk_found = True
                        if row_count > 2:
                            ddl += beauty_empty(N_COLUMN,column_nm,column_type)
                            if column_comment != 'None':
                                comment_text += COMMENT.replace('[RAPLACE_COLUMN_NM]',column_nm).replace('[RAPLACE_TABLE_NM]',table_nm).replace('[RAPLACE_COMMENT]',column_comment)
                            if index_ind == '*':
                                index += INDEX.replace('[RAPLACE_TABLE_NM]',table_nm).replace('[RAPLACE_COLUMN_NM]',column_nm)
                            if pk_ind == 'PK':
                                pk_found = True
                                pk_column += column_nm + ','
                if table_found:
                    ddl += '''		CONSTRAINT PK_[RAPLACE_TABLE_NM] PRIMARY KEY CLUSTERED (R_PK_COLUMN)
    )ON [{TABLEFG}]\n\n'''.replace('[RAPLACE_TABLE_NM]',table_nm).replace('R_PK_COLUMN',pk_column[1:])
                    ddl += comment_text
                    ddl += '''\n	PRINT '[INFO] CREATED TABLE [DBO].[[RAPLACE_TABLE_NM]]'
END\nGO\n\n'''.replace('[RAPLACE_TABLE_NM]',table_nm)
                    ddl += index
                    ddl += 'GO'
            if table_found and not pk_found:
                print(table_nm+' does not have PK column.')

        save_file(r'.\output\ddl\\'+table_nm+'.sql',ddl)



def generate_oracle_ddl():

    output = ''
    for table_nm in gen_tables:

        ddl = ''
        comment_text = ''
        row_count = 0
        pk_column = ''
        index = ''
        constraint = ''
        
        for sheetname in sheetnames:
            table_found = False
            pk_found = False
            if sheetname != 'ChangeLog':
                sheet = workbook.get_sheet_by_name(sheetname)
                for row in range(1,sheet.max_row+1):
                    if sheet.cell(row=row,column=3).value == table_nm:
                        table_found = True
                        ms_column_nm = str(sheet.cell(row=row,column=4).value)
                        oracle_column_nm = str(sheet.cell(row=row,column=22).value)
                        ms_column_type = str(sheet.cell(row=row,column=6).value)
                        oracle_column_type = str(sheet.cell(row=row,column=23).value)
                        pk_ind = str(sheet.cell(row=row,column=7).value)
                        column_comment = str(sheet.cell(row=row,column=19).value)
                        oracle_table_nm = str(sheet.cell(row=row,column=21).value)
                        row_count +=1
                        output_table_nm = oracle_table_nm if oracle_table_nm != 'None' else table_nm

                        if row_count == 1:
                            ddl += ORACLE_HEADER.replace('[RAPLACE_TABLE_NM]',output_table_nm)
                            continue

                        if oracle_column_nm == 'None' and len(ms_column_nm) > 30:
                            print("\n\033[31m" + table_nm + '\033[0m.\033[33m' + ms_column_nm + "\033[0m is longer than 30")
                        
                        if oracle_table_nm == 'None' and len(table_nm) > 27:
                            print("\n\033[31m" + table_nm + "\033[0m name is longer than 27")

                        column_nm = oracle_column_nm if oracle_column_nm != 'None' else ms_column_nm
                        column_type = oracle_column_type if oracle_column_type != 'None' else ms_column_type

                        if pk_ind == 'PK':
                            pk_found = True
                            pk_column += column_nm + ','

                        ddl += beauty_empty_oracle(ORACLE_COLUMN,column_nm,column_type)
                        if column_comment != 'None':
                            comment_text += ORACLE_COMMENT.replace('[RAPLACE_COLUMN_NM]',column_nm).replace('[RAPLACE_TABLE_NM]',output_table_nm).replace('[RAPLACE_COMMENT]',column_comment.replace("'","''"))

                if pk_found:

                    index += ORACLE_PK_INDEX.replace('[RAPLACE_TABLE_NM]',output_table_nm).replace('[RAPLACE_PK_NM]',pk_column[:len(pk_column)-1])
                    constraint += ORACLE_PK_CONSTRAINT.replace('[RAPLACE_TABLE_NM]',output_table_nm).replace('[RAPLACE_PK_NM]',pk_column[:len(pk_column)-1])

                    '''if column_nm == 'MART_SOURCE_ID':
                        has_mart_source_id = True
                    if column_nm.endswith('_KEY'):
                        index += INDEX.replace('[RAPLACE_TABLE_NM]',table_nm).replace('[RAPLACE_COLUMN_NM]',column_nm)
                    '''
                if table_found and not pk_found:
                    print(table_nm+' does not have PK column.')
                if table_found:
                    ddl = ddl[:len(ddl)-2]+'''\n) TABLESPACE {TABLESPACE}
;\n\n'''
                    ddl += comment_text
                    ddl += index
                    ddl += constraint
        output += ddl
    save_file(r'.\output\ddl\oracle.sql',output)

def check_name(name, type):

    incorrect_ind = False
    new_name = ''

    spell = SpellChecker()

    full_name = spell.split_words(name.replace('_',' '))
    for word in full_name:
        if (full_name[0] == word and word in ('d','b','f','r')) or spell.correction(word) == word:
            new_name += word + ' '
        elif (full_name[-1] == word and word in ('key','txt','nb','amt','dtm','qty','dt')) or word in ('prev'):
            new_name += word + ' '
        else:
            incorrect_ind = True
            new_name += spell.correction(word) + ' '

    if type != None:
        if full_name[-1] in ('key','id') and type not in ('smallint','int','bigint'):
            print('column ' + name + ' type: <\028[32m' + type + '\033[0m> may be wrong.')
        elif full_name[-1] == 'dtm' and type not in ('datetime'):
            print('column ' + name + ' type: <\028[32m' + type + '\033[0m> may be wrong.')
        elif full_name[-1] == 'nm' and not (str(type).startswith('varchar') or str(type).startswith('char')):
            print('column ' + name + ' type: <\028[32m' + type + '\033[0m> may be wrong.')
        elif full_name[-1] == 'txt' and not (str(type).startswith('varchar') or str(type).startswith('char')):
            print('column ' + name + ' type: <\028[32m' + type + '\033[0m> may be wrong.')

    if incorrect_ind:
        print(name + ' should be \033[32m' + new_name.upper().replace(' ','_')[:len(new_name)-1] +'\033[0m')


generate_ms_ddl()
#generate_oracle_ddl()
