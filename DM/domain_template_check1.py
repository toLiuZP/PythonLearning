###
# 1. Verify if C_CUST_HFPROFILE has data. Impact D_CUSTOMER and D_CUSTOMER_ADDRESS
# 2. Confirm fiscal date.
# 3. Verify what occupant type this contract has.
# 	select * from P_ADMISSION_type where id in(
# 		select ADMISSION_type_id from P_ADMISSION_PRD_CAT
# 		)		
# 		order by 1; 
# 4. Verify ticket types
# 	select * from P_ADMISSION_type where id in(
# 		select ADMISSION_type_id from O_TICKET_QUANTITY
# 		)		
# 		order by 1; 
# 5. Verify site attributes.
# 	SELECT 
#         DISTINCT
#         a.attr_id
#         ,a.attr_name
#     FROM 
#         p_prd p 
#         LEFT JOIN
#         p_prd pp on pp.prd_id = p.parent_id and p.prd_rel_type = 3 -- Child
#         LEFT JOIN 
#         p_prd_attr pa ON pa.prd_id = COALESCE( pp.prd_id, p.prd_id )
#         LEFT JOIN 
#         d_attr a ON a.attr_id = pa.attr_id
#     WHERE 
#         p.product_cat_id = 3 -- Site
#         AND
#         pa.active_ind = 1
#         AND
#         pa.deleted_ind = 0
# 		order by a.attr_name
###


import numpy as np
import pandas as pd
#import cx_Oracle as oracle
import os

from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter
import time

#import conf.acct_oracle as acct_oracle
#from db_connect.oracle_db import UseOracleDB
#from tool.df_compare import has_gap

#CURRENT_DB = acct_oracle.QA3
SCHEMA = 'LIVE_CO'
#SEED_FILE = '.\seed\Domain Data Template.xlsx'
SEED_FILE = 'DDT.xlsx'

writer = pd.ExcelWriter('est.xlsx')

os.system("")
nameTime = time.strftime('%Y%m%d_%H%M%S')
excelName = 'est' + nameTime + '.xlsx'

workbook_ = load_workbook(SEED_FILE)

sheetnames =workbook_.get_sheet_names() 

query_sheet = workbook_.get_sheet_by_name('Checking_Query')

for sheetname in sheetnames:
    sheet = workbook_.get_sheet_by_name(sheetname)

    if sheetname != 'Checking_Query':
        rows = sheet.rows
        columns = sheet.columns

        if sheet['A1'].value == 'Datasets':
            for col in range(2, sheet.max_column+1, 2):
                for row in range(3,sheet.max_row+1):
                    domain = sheetname
                    dataset = sheet.cell(row=row,column=1).value

                    for query_row in range(1,query_sheet.max_row+1):
                        if domain == query_sheet.cell(row=query_row,column=1).value and dataset == query_sheet.cell(row=query_row,column=2).value:
                            query_txt = query_sheet.cell(row=query_row,column=3).value
                            if str(query_txt).strip() != 'None':
                                schema = "LIVE_" + str(sheet.cell(row=1,column=col).value)
                                query_txt = str(query_txt).replace('{SOURCE_SCHEMA}',schema)
                                print(query_txt)
                                sheet.cell(row=row,column=col).value = 'X'
        elif sheet['A1'].value == 'Domain\nFields':

            for i in range(3, sheet.max_column+1, 2):
                for j in range(3,sheet.max_row+1):
                    sheet.cell(row=j,column=i).value = 'X'
    '''

    for row in rows:

        for col in row:
            print("test")

        #line = [col.value for col in row]

        #print(line)

    print(sheet.cell(row=3,column=3).value)
    '''

#sheet['A1'] = '47' 

workbook_.save(excelName)  

print("test")
'''
wb = Workbook()

ws = wb.active

ws['A1'] = 4

wb.save("新歌检索失败.xlsx") 
'''
'''
test_e = pd.read_excel(SEED_FILE)
df = pd.DataFrame(test_e)

df.index = df.Datasets
df.loc['Customer_Profile','KS'] = 'X'

df.to_excel(writer,sheet_name = '123')
writer.save()

pass
pass
'''
'''
with UseOracleDB(CURRENT_DB) as cursor:

    # 1. Verify if C_CUST_HFPROFILE has data. Impact D_CUSTOMER and D_CUSTOMER_ADDRESS
    has_customer_hfprofile = "SELECT * FROM " + SCHEMA + ".C_CUST_HFPROFILE WHERE ID > 1 AND ROWNUM < 2"
    cursor.execute(has_customer_hfprofile)
    row = cursor.fetchall()
    if len(row) == 0:
        print("C_CUST_HFPROFILE is empty. Please commentet customer number and birthday")
    else:
        print("Please use hfprofile and load customer number and birthday")

    # 3. Verify what occupant type this contract has.

    inquery_occupant_sql = "select ID, NAME from " + SCHEMA + ".P_ADMISSION_TYPE where id in (select ADMISSION_TYPE_ID from " + SCHEMA + ".P_ADMISSION_PRD_CAT) order by 1"

    cursor.execute(inquery_occupant_sql)
    row = cursor.fetchall()

    if len(row) == 0:
        print(SCHEMA + "'s occupant type is null.")
    else:
        occupant_type_seed = pd.read_excel(SEED_FILE,sheet_name = "P_ADMISSION_PRD_CAT")
        occupant_type_return = pd.DataFrame(row)
        has_gap(occupant_type_seed,occupant_type_return,"Occupant type")

    # 4. Verify ticket types

    inquery_ticket_sql = "select ID, NAME from " + SCHEMA + ".P_ADMISSION_TYPE where id in (select ADMISSION_TYPE_ID from " + SCHEMA + ".O_TICKET_QUANTITY) order by 1"

    cursor.execute(inquery_ticket_sql)
    row = cursor.fetchall()

    if len(row) == 0:
        print(SCHEMA + "'s ticket type is null.")
    else:        
        ticket_type_seed = pd.read_excel(SEED_FILE,sheet_name = "O_TICKET_QUANTITY")
        ticket_type_return = pd.DataFrame(row)
        has_gap(ticket_type_seed,ticket_type_return,"ticket type")

    # 5. Verify site attributes.

    inquery_site_attr_sql = "SELECT DISTINCT a.attr_id AS ID, a.attr_name FROM " + SCHEMA + ".p_prd p LEFT JOIN " + SCHEMA + ".p_prd pp on pp.prd_id = p.parent_id and p.prd_rel_type = 3 LEFT JOIN " + SCHEMA + ".p_prd_attr pa ON pa.prd_id = NVL( pp.prd_id, p.prd_id ) LEFT JOIN " + SCHEMA + ".d_attr a ON a.attr_id = pa.attr_id WHERE p.product_cat_id = 3 AND pa.active_ind = 1 AND pa.deleted_ind = 0 order by a.attr_id, a.attr_name"

    cursor.execute(inquery_site_attr_sql)
    row = cursor.fetchall()

    if len(row) == 0:
        print(SCHEMA + "'s site attributes is null.")
    else:
        site_attr_seed = pd.read_excel(SEED_FILE,sheet_name = "SITE_ATTRIBUTES")
        site_attr_return = pd.DataFrame(row)
        has_gap(site_attr_seed,site_attr_return,"Site Attributes")
'''
    
    



    
        

