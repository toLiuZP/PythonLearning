######
# TODO: result shoudl sort by column_seq
#
######



import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
import openpyxl.styles as sty


import conf.acct as acct
import db_connect.db_operator as DB
from tool.tool import file_name,logger,identify_backup_tables

SEED_FILE = r".\seed\DDL_GAP_AB.xlsx"
excelName = file_name('DDL_GAP_AB','xlsx')
workbook = load_workbook(SEED_FILE)
ddl_sheet = workbook.get_sheet_by_name('DDL')
sp_sheet = workbook.get_sheet_by_name('SP')
index_sheet = workbook.get_sheet_by_name('INDEX')


def merge_ddl(db1, db2, sheet):

    db1 = db1.fillna('null')
    db2 = db2.fillna('null')

    gap = pd.merge(db1, db2, on = ['table_name','column_name'], how='outer')

    for index, col in gap.iterrows():
        if identify_backup_tables(col[0].lower()):
            gap = gap.drop(index)
            continue
        
        #print(col[0]+":"+col[1])
        if col[3] == col[8] and col[4] == col[9] \
            and col[5] == col[10] and col[6] == col[11] :
            '''Don't compare postion col[2] == col[7]  '''
            gap = gap.drop(index)

    gap = gap.sort_values(by = ['table_name','column_name'])

    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+3,column=i+1).value = col[i]
            if i in (2,3,4,5,6):
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i in (7,8,9,10,11):   
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")


def merge_sp(db1, db2, sheet):

    db1 = db1.fillna('null')
    db2 = db2.fillna('null')
    gap = pd.merge(db1, db2, on = ['Type','name'], how='outer')

    for index, col in gap.iterrows():
        if identify_backup_tables(str(col[0]).lower()):
            gap = gap.drop(index)
        
        sp_a = str(col[2]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','').upper()
        sp_b = str(col[3]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','').upper()
        if sp_a == sp_b:
            gap = gap.drop(index)

    gap = gap.fillna('null')
    gap = gap.reset_index(drop=True)

    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+2,column=i+1).value = col[i]
            if i == 2:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i == 3:   
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")
            elif i == 4:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="68AE59")
           

def merge_index(db1, db2, sheet):

    db1 = db1.fillna('null')
    db2 = db2.fillna('null')

    gap = pd.merge(db1, db2, on = ['table_nm','index_nm'], how='outer')

    for index, col in gap.iterrows():
        if identify_backup_tables(col[0].lower()):
            gap = gap.drop(index)
            continue
        if col[2] == col[3]:
            gap = gap.drop(index)
                
    gap = gap.fillna('null')

    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+2,column=i+1).value = col[i]
            if i == 2:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i == 3:   
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")

def check_ddl(workbook,ddl_sheet,db_a, db_b):

    query = "SELECT table_name, column_name, ordinal_position , data_type, COALESCE(character_maximum_length,numeric_precision,datetime_precision) data_length, numeric_scale, is_nullable FROM information_schema.columns WHERE table_schema = 'dbo' AND (table_name LIKE 'D[_]%' OR table_name LIKE 'B[_]%' OR table_name LIKE 'R[_]%' OR table_name LIKE 'F[_]%' OR table_name LIKE 'RPT[_]%') ORDER BY table_name"

    new_sheet = workbook.copy_worksheet(ddl_sheet)
    new_sheet.title = "DDL"
    ddl_a = DB.query_db_pandas(query, db_a)
    ddl_b = DB.query_db_pandas(query, db_b)

    merge_ddl(ddl_a,ddl_b,new_sheet)
    

def check_sp(workbook,sp_sheet,db_a, db_b):

    query = "SELECT CASE a.[type] WHEN 'P' THEN 'Stored Procedures' WHEN 'V' THEN 'Views' WHEN 'AF' THEN 'Aggregate function' END AS 'Type', a.name, b.[definition] FROM sys.all_objects a, sys.sql_modules b WHERE a.is_ms_shipped=0 AND a.object_id = b.object_id AND a.[type] IN ('P','V','AF') order by a.[type], a.[name] ASC"

    new_sheet = workbook.copy_worksheet(sp_sheet)
    new_sheet.title = "SP"
    sp_a = DB.query_db_pandas(query, db_a)
    sp_b = DB.query_db_pandas(query, db_b)

    merge_sp(sp_a,sp_b,new_sheet)


def check_index(workbook,ddl_sheet,db_a, db_b):

    query = "SELECT DISTINCT o.name as table_nm,i.name as index_nm, 'Y' as val FROM SYS.OBJECTS O JOIN SYS.index_columns IC ON IC.OBJECT_ID = O.OBJECT_ID JOIN SYS.COLUMNS C ON IC.column_id = C.column_id and C.OBJECT_ID = O.OBJECT_ID JOIN SYS.INDEXES I ON I.OBJECT_ID = O.OBJECT_ID AND I.index_id = IC.index_id JOIN information_schema.tables tab ON O.NAME = tab.TABLE_NAME AND tab.table_schema = 'dbo' WHERE o.type = 'U' AND (table_name LIKE 'D[_]%' OR table_name LIKE 'B[_]%' OR table_name LIKE 'R[_]%' OR table_name LIKE 'F[_]%' OR table_name LIKE 'RPT[_]%') ORDER BY o.name,i.name"

    new_sheet = workbook.copy_worksheet(ddl_sheet)
    new_sheet.title = "INDEX"
    index_a = DB.query_db_pandas(query, db_a)
    index_b = DB.query_db_pandas(query, db_b)

    merge_index(index_a,index_b,new_sheet)


if __name__ == '__main__':

    db_a = acct.DEV_DMA_MART_TEST
    db_b = acct.QA_OR_CAMPING_MART

    check_ddl(workbook,ddl_sheet,db_a,db_b)
    check_sp(workbook,sp_sheet,db_a,db_b)
    check_index(workbook,index_sheet,db_a,db_b)
    
workbook.remove_sheet(workbook.get_sheet_by_name('DDL'))
workbook.remove_sheet(workbook.get_sheet_by_name('SP'))
workbook.remove_sheet(workbook.get_sheet_by_name('INDEX'))
workbook.save(excelName)
