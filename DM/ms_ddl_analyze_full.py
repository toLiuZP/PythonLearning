###
#
###

import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
import openpyxl.styles as sty

import conf.acct as acct
import db_connect.db_operator as DB
from tool.tool import file_name,logger,identify_backup_tables

SEED_FILE = r".\seed\DDL_GAP.xlsx"
excelName = file_name('DDL_GAP','xlsx')
workbook = load_workbook(SEED_FILE)
ddl_sheet = workbook.get_sheet_by_name('DDL')
view_sheet = workbook.get_sheet_by_name('VIEW')
view_sheet = workbook.get_sheet_by_name('INDEX')

#check_list = ['CO_HF_MART','KS_HF_MART','MS_HF_MART','TX_CAMPING_MART']
check_list = ['TX_CAMPING_MART']

def merge_ddl(dev, qa, uat, sheet, prod='none'):

    dev = dev.fillna('null')
    qa = qa.fillna('null')
    uat = uat.fillna('null')

    gap = pd.merge(dev, qa, on = ['table_name','column_name'], how='outer')
    gap = pd.merge(gap, uat, on = ['table_name','column_name'], how='outer')

    if isinstance(prod, str):
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            if col[3] == col[8] == col[13] and col[4] == col[9] == col[14] \
                and col[5] == col[10] == col[15] and col[6] == col[11] == col[16]:
                '''Don't compare postion col[2] == col[7] == col[12] and '''
                gap = gap.drop(index)
                
    else:
        prod = prod.fillna('null')
        gap = pd.merge(gap, prod, on = ['table_name','column_name'], how='outer')
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            if  col[3] == col[8] == col[13]  == col[18] and col[4] == col[9] == col[14]  == col[19] \
                and col[5] == col[10] == col[15]  == col[20] and col[6] == col[11] == col[16]  == col[21]:
                '''Don't compare postion col[2] == col[7] == col[12] == col[17] and'''
                gap = gap.drop(index)
    
    gap = gap.sort_values(by = ['table_name','column_name'])

    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+3,column=i+1).value = col[i]
            if i in (2,3,4,5,6):
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i in (7,8,9,10,11):   
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")
            elif i in (12,13,14,15,16):
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="68AE59")
            elif i in (17,18,19,20,21):
                sheet.cell(row=index+3,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="5B7190")

def merge_view(dev, qa, uat, sheet, prod='none'):

    dev = dev.fillna('null')
    qa = qa.fillna('null')
    uat = uat.fillna('null')

    gap = pd.merge(dev, qa, on = ['Type','name'], how='outer')
    gap = pd.merge(gap, uat, on = ['Type','name'], how='outer')

    if isinstance(prod, str):
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            col[2] = str(col[2]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            col[3] = str(col[3]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            col[4] = str(col[4]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            if col[2] == col[3] == col[4]:
                gap = gap.drop(index)
                
    else:
        prod = prod.fillna('null')
        gap = pd.merge(gap, prod, on = ['Type','name'], how='outer')
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            col2 = str(col[2]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            col3 = str(col[3]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            col4 = str(col[4]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            col5 = str(col[5]).replace('[','').replace(']','').replace('\r','').replace('\t','').replace('\n','').replace(' ','')
            if  col2 == col3 == col4 == col5:
                gap = gap.drop(index)

    gap = gap.fillna('null')

    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+2,column=i+1).value = col[i]
            if i == 2:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i == 3:   
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")
            elif i == 4:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="68AE59")
            elif i == 5:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="5B7190")

def merge_index(dev, qa, uat, sheet, prod='none'):

    dev = dev.fillna('null')
    qa = qa.fillna('null')
    uat = uat.fillna('null')

    gap = pd.merge(dev, qa, on = ['table_nm','index_nm'], how='outer')
    gap = pd.merge(gap, uat, on = ['table_nm','index_nm'], how='outer')

    if isinstance(prod, str):
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            if col[2] == col[3] == col[4]:
                gap = gap.drop(index)
                
    else:
        prod = prod.fillna('null')
        gap = pd.merge(gap, prod, on = ['table_nm','index_nm'], how='outer')
        for index, col in gap.iterrows():
            if identify_backup_tables(col[0].lower()):
                gap = gap.drop(index)
            if  col[2] == col[3] == col[4] == col[5]:
                gap = gap.drop(index)

    gap = gap.fillna('null')

    gap = gap.reset_index(drop=True)
    for index, col in gap.iterrows():
        for i in range(0, len(col), 1):
            sheet.cell(row=index+2,column=i+1).value = col[i]
            if i == 2:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="C28EEA")
            elif i == 3:   
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="DAC45E")
            elif i == 4:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="68AE59")
            elif i == 5:
                sheet.cell(row=index+2,column=i+1).fill=sty.PatternFill(fill_type='solid',fgColor="5B7190")

def check_ddl(workbook,ddl_sheet,contract):

    print("checking " + contract + " ddl")

    query = "SELECT table_name, column_name, ordinal_position , data_type, COALESCE(character_maximum_length,numeric_precision,datetime_precision) data_length, numeric_scale, is_nullable FROM " + contract + ".information_schema.columns WHERE table_schema = 'dbo' AND (table_name LIKE 'D[_]%' OR table_name LIKE 'B[_]%' OR table_name LIKE 'R[_]%' OR table_name LIKE 'F[_]%' OR table_name LIKE 'RPT[_]%') ORDER BY table_name"

    new_sheet = workbook.copy_worksheet(ddl_sheet)
    new_sheet.title = contract + "_DDL"
    dev = DB.query_db_pandas(query, acct.DEV_CO_HF_MART)
    qa = DB.query_db_pandas(query, acct.QA_CO_HF_MART)
    uat = DB.query_db_pandas(query, acct.UAT_CO_HF_MART)
    prod = DB.query_db_pandas(query, acct.PROD_CO_HF_MART)

    #merge_ddl(dev,qa,uat,new_sheet)
    merge_ddl(dev,qa,uat,new_sheet,prod)

def check_view(workbook,ddl_sheet,contract):

    print("checking " + contract + " view")

    query = "SELECT CASE a.[type] WHEN 'P' THEN 'Stored Procedures' WHEN 'V' THEN 'Views' WHEN 'AF' THEN 'Aggregate function' END AS 'Type', a.name, b.[definition] FROM " + contract + ".sys.all_objects a, " + contract + ".sys.sql_modules b WHERE a.is_ms_shipped=0 AND a.object_id = b.object_id AND a.[type] IN ('P','V','AF') order by a.[type], a.[name] ASC"

    new_sheet = workbook.copy_worksheet(ddl_sheet)
    new_sheet.title = contract + "_VIEW"
    dev = DB.query_db_pandas(query, acct.DEV_CO_HF_MART)
    qa = DB.query_db_pandas(query, acct.QA_CO_HF_MART)
    uat = DB.query_db_pandas(query, acct.UAT_CO_HF_MART)
    prod = DB.query_db_pandas(query, acct.PROD_CO_HF_MART)

    #merge_view(dev,qa,uat,new_sheet)
    merge_view(dev,qa,uat,new_sheet,prod)

def check_index(workbook,ddl_sheet,contract):

    print("checking " + contract + " index")

    query = "SELECT DISTINCT o.name as table_nm,i.name as index_nm, 'Y' as val FROM " + contract + ".SYS.OBJECTS O JOIN " + contract + ".SYS.index_columns IC ON IC.OBJECT_ID = O.OBJECT_ID JOIN " + contract + ".SYS.COLUMNS C ON IC.column_id = C.column_id and C.OBJECT_ID = O.OBJECT_ID JOIN " + contract + ".SYS.INDEXES I ON I.OBJECT_ID = O.OBJECT_ID AND I.index_id = IC.index_id JOIN " + contract + ".information_schema.tables tab ON O.NAME = tab.TABLE_NAME AND tab.table_schema = 'dbo' WHERE o.type = 'U' AND (table_name LIKE 'D[_]%' OR table_name LIKE 'B[_]%' OR table_name LIKE 'R[_]%' OR table_name LIKE 'F[_]%' OR table_name LIKE 'RPT[_]%') ORDER BY o.name,i.name"

    new_sheet = workbook.copy_worksheet(ddl_sheet)
    new_sheet.title = contract + "_INDEX"
    dev = DB.query_db_pandas(query, acct.DEV_CO_HF_MART)
    qa = DB.query_db_pandas(query, acct.QA_CO_HF_MART)
    uat = DB.query_db_pandas(query, acct.UAT_CO_HF_MART)
    prod = DB.query_db_pandas(query, acct.PROD_CO_HF_MART)

    #merge_index(dev,qa,uat,new_sheet)
    merge_index(dev,qa,uat,new_sheet,prod)


if __name__ == '__main__':

    # US contracts:
    for contract in check_list:
        check_ddl(workbook,ddl_sheet,contract)
        #check_view(workbook,view_sheet,contract)
        check_index(workbook,view_sheet,contract)
    
workbook.remove_sheet(workbook.get_sheet_by_name('DDL'))
workbook.remove_sheet(workbook.get_sheet_by_name('VIEW'))
workbook.remove_sheet(workbook.get_sheet_by_name('INDEX'))
workbook.save(excelName)
