###
# 1. Verify if C_CUST_HFPROFILE has data. Impact D_CUSTOMER and D_CUSTOMER_ADDRESS
# 2. Confirm fiscal date.
# 3. Verify what occupant type this contract has.
# 	select * from P_ADMISSION_type where id in(
# 		select ADMISSION_type_id from P_ADMISSION_PRD_CAT
# 		)		
# 		order by 1; 
# 4. Verify ticket types
# 	select * from P_ADMISSION_type where id in(
# 		select ADMISSION_type_id from O_TICKET_QUANTITY
# 		)		
# 		order by 1; 
# 5. Verify site attributes.
# 	SELECT 
#         DISTINCT
#         a.attr_id
#         ,a.attr_name
#     FROM 
#         p_prd p 
#         LEFT JOIN
#         p_prd pp on pp.prd_id = p.parent_id and p.prd_rel_type = 3 -- Child
#         LEFT JOIN 
#         p_prd_attr pa ON pa.prd_id = COALESCE( pp.prd_id, p.prd_id )
#         LEFT JOIN 
#         d_attr a ON a.attr_id = pa.attr_id
#     WHERE 
#         p.product_cat_id = 3 -- Site
#         AND
#         pa.active_ind = 1
#         AND
#         pa.deleted_ind = 0
# 		order by a.attr_name
###

import numpy as np
import pandas as pd
import cx_Oracle as oracle
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time

import conf.acct_oracle as acct_oracle
from db_connect.oracle_db import UseOracleDB
from tool.df_compare import has_gap
import tool.oracle_tool as oracle_tool
import tool.tool as tool

CURRENT_DB = acct_oracle.PROD_US
os.system("")

SEED_FILE = ".\seed\Domain Data Template.xlsx"
nameTime = time.strftime('%Y%m%d_%H%M%S')
excelName = tool.file_name('Domain Data Template','xlsx')
workbook = load_workbook(SEED_FILE)
sheetnames =workbook.get_sheet_names() 
query_sheet = workbook.get_sheet_by_name('Checking_Query')

check_list = []

with UseOracleDB(CURRENT_DB) as cursor:

    for sheetname in sheetnames:
        sheet = workbook.get_sheet_by_name(sheetname)

        if sheetname != 'Checking_Query':
            rows = sheet.rows
            columns = sheet.columns

            if sheet['A1'].value == 'Datasets':
                for col in range(2, sheet.max_column+1, 2):
                    if (len(check_list) == 0) or (len(check_list) > 0 and sheet.cell(row=1,column=col).value in check_list):
                        for row in range(3,sheet.max_row+1):
                            domain = sheetname
                            dataset = sheet.cell(row=row,column=1).value

                            for query_row in range(1,query_sheet.max_row+1):
                                if domain == query_sheet.cell(row=query_row,column=1).value and dataset == query_sheet.cell(row=query_row,column=2).value:
                                    query_txt = query_sheet.cell(row=query_row,column=3).value
                                    if str(query_txt).strip() != 'None':
                                        schema = "LIVE_" + str(sheet.cell(row=1,column=col).value)
                                        query_txt = str(query_txt).replace('{SOURCE_SCHEMA}',schema)
                                        has_row = oracle_tool.has_row(query_txt,cursor)
                                        if has_row:
                                            sheet.cell(row=row,column=col).value = 'X'
                                if sheetname == 'Customer Activity' and sheet.cell(row=row,column=1).value in ('User Created','User Modified'):
                                    sheet.cell(row=row,column=col).value = 'X'
                                if sheetname == 'Financial' and sheet.cell(row=row,column=1).value in ('Payment User'):
                                    sheet.cell(row=row,column=col).value = 'X'

            elif sheet['A1'].value == 'Domain\nFields':
                for i in range(3, sheet.max_column+1, 2):
                    if (len(check_list) == 0) or (len(check_list) > 0 and sheet.cell(row=1,column=i).value in check_list):
                        schema = "LIVE_" + str(sheet.cell(row=1,column=i).value)
                        for j in range(3,sheet.max_row+1):
                            if sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value in ('Birth Date','Customer Number'):
                                query = 'SELECT CUST_ID FROM {SOURCE_SCHEMA}.c_cust_hfprofile WHERE CUST_ID > 1 AND ROWNUM = 1'.replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value in ('First Name','Last Name','Middle Name','Salutation','Suffix'):
                                query = 'SELECT CUST_ID FROM {SOURCE_SCHEMA}.c_cust WHERE ROWNUM = 1'.replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value == 'Customer Deleted':
                                sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value == 'Home Phone Number':
                                query = "SELECT cust_id FROM {SOURCE_SCHEMA}.c_cust_phone WHERE typ = 'HOME' AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value == 'Work Phone Number':
                                query = "SELECT cust_id FROM {SOURCE_SCHEMA}.c_cust_phone WHERE typ = 'WORK' AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value == 'Mobile Phone Number':
                                query = "SELECT cust_id FROM {SOURCE_SCHEMA}.c_cust_phone WHERE typ = 'CELL' AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Customer_Profile' and sheet.cell(row=j,column=1).value == 'Email':
                                query = "SELECT cust_id FROM {SOURCE_SCHEMA}.c_cust_phone WHERE typ = 'EMAIL' AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'Agency':
                                query = "SELECT name_20 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_20 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'District':
                                query = "SELECT name_32 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_32 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'Facility':
                                query = "SELECT name_40 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_40 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'Region':
                                query = "SELECT name_30 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_30 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'Project':
                                query = "SELECT name_35 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_35 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value == 'Facility HQ':
                                query = "SELECT name_39 FROM {SOURCE_SCHEMA}.D_LOC_HIERARCHY WHERE name_39 IS NOT NULL AND ROWNUM = 1".replace('{SOURCE_SCHEMA}',schema)
                                if oracle_tool.has_row(query,cursor):
                                    sheet.cell(row=j,column=i).value = 'X'
                            elif sheetname == 'Location' and sheet.cell(row=j,column=1).value in ('Location Category','Location Deleted'):
                                sheet.cell(row=j,column=i).value = 'X'
workbook.save(excelName)  
